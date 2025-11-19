import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import face_recognition
import cv2
import os
import numpy as np
import openpyxl
from datetime import datetime, timedelta
import threading
from flask import Flask, request
from werkzeug.serving import make_server
import queue
import time

# ========== CONFIG ==========
USER_FILE = "authorized_users.csv"
EXCEL_FILE = "Book3.xlsx"
KNOWN_FACES_DIR = "known_faces"
HOST = "0.0.0.0"
PORT = 5000
ATTENDANCE_DURATION = 50*60 

# ========== GLOBALS ==========
known_face_encodings = []
known_face_names = []
card_to_name = {}
app = Flask(__name__)
server_thread = None

# Queue for safe communication between Flask thread and Tkinter main thread
request_queue = queue.Queue()
pending_results = {}  # card_id -> {"event": Event, "result": str}
attendance_tracking = {}  # card_id -> {"name": str, "entry_time": datetime, "marked": bool}

# ========== GUI SETUP ==========
root = tk.Tk()
root.title("RFID + Face Recognition Attendance (Wi-Fi)")
root.geometry("900x650")
root.configure(bg="#7c9997")

style = ttk.Style()
style.configure("TButton", font=("Segoe UI", 12), padding=6)
style.configure("TLabel", font=("Segoe UI", 12))
style.configure("TEntry", font=("Segoe UI", 12))

log_text = tk.Text(root, height=15, width=105, font=("Courier New", 10))
log_text.pack(pady=10)

canvas = tk.Canvas(root, width=200, height=200, bg="#80b1a5", bd=2, relief="groove")
canvas.pack(pady=10)

status_label = tk.Label(root, text="", font=("Segoe UI", 14, "bold"), bg="#7ca89d")
status_label.pack()

# ========== UTILS ==========
def append_log(entry):
    log_text.insert(tk.END, entry + "\n")
    log_text.see(tk.END)

def load_faces():
    known_face_encodings.clear()
    known_face_names.clear()
    if not os.path.exists(KNOWN_FACES_DIR):
        os.makedirs(KNOWN_FACES_DIR)
    for filename in os.listdir(KNOWN_FACES_DIR):
        if filename.endswith(".jpg"):
            path = os.path.join(KNOWN_FACES_DIR, filename)
            img = face_recognition.load_image_file(path)
            enc = face_recognition.face_encodings(img)
            if enc:
                known_face_encodings.append(enc[0])
                known_face_names.append(filename.split(".")[0].lower())  
            else:
                print(f"[WARN] No face in {KNOWN_FACES_DIR}")

def load_user_records():
    global card_to_name
    card_to_name = {}
    if os.path.exists(USER_FILE):
        with open(USER_FILE, "r") as f:
            for line in f:
                parts = line.strip().split(",")
                if len(parts) == 2:
                    card_id, name = parts
                    card_to_name[card_id] = name.lower()  

def save_to_excel(name, card_id, status, attendance=None):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    sheet = wb.active
    if sheet.max_row == 1:
        sheet.append(["Timestamp", "Card ID", "Name", "Status", "Attendance"])
    sheet.append([now, card_id, name, status, attendance if attendance else ""])
    wb.save(EXCEL_FILE)

def update_attendance_excel(card_id, attendance_status):
    """Mark Present/Absent for a user in Excel after 50 minutes."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    sheet = wb.active
    name = attendance_tracking[card_id]["name"]
    sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), card_id, name, "Authorized", attendance_status])
    wb.save(EXCEL_FILE)
    append_log(f"[ATTENDANCE] {name} marked as {attendance_status}")

def mark_present_after_time(card_id, duration_seconds):
    """Wait duration, then mark Present."""
    time.sleep(duration_seconds)
    if card_id in attendance_tracking and not attendance_tracking[card_id]["marked"]:
        attendance_tracking[card_id]["marked"] = True
        update_attendance_excel(card_id, "Present")

def show_image(img_path):
    img = Image.open(img_path)
    img = img.resize((200, 200))
    img_tk = ImageTk.PhotoImage(img)
    canvas.create_image(0, 0, anchor=tk.NW, image=img_tk)
    canvas.image = img_tk

def rotate_image(image, angle):
    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    matrix = cv2.getRotationMatrix2D(center, angle, 1.0)
    return cv2.warpAffine(image, matrix, (w, h))

# ========== FACE HANDLER ==========
def handle_card(card_id):
    if card_id not in card_to_name:
        append_log("❌ Unauthorized card")
        save_to_excel("Unknown", card_id, "Unauthorized", "Absent")
        status_label.config(text="Unauthorized Card", fg="red")
        return "UNAUTH"

    expected_name = card_to_name[card_id]
    append_log(f"[INFO] Expected: {expected_name}")

    # Track attendance
    if card_id not in attendance_tracking:
        attendance_tracking[card_id] = {"name": expected_name, "entry_time": datetime.now(), "marked": False}
        threading.Thread(target=mark_present_after_time, args=(card_id, ATTENDANCE_DURATION), daemon=True).start()

    cap = cv2.VideoCapture(0)
    matched = False
    matched_name = "Unknown"
    frame_to_show = None
    best_distance = float('inf')

    for attempt in range(3):
        ret, frame = cap.read()
        if not ret:
            append_log(f"[DEBUG] Attempt {attempt+1}: Failed to capture frame")
            continue
        rgb_frame = frame[:, :, ::-1]

        for angle in [0, -15, 15, -30, 30]:
            rotated = rotate_image(rgb_frame, angle)
            face_locations = face_recognition.face_locations(rotated)
            face_encodings = face_recognition.face_encodings(rotated, face_locations)

            for face_encoding in face_encodings:
                distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                if len(distances) == 0:
                    continue
                min_distance = np.min(distances)
                best_idx = np.argmin(distances)
                if min_distance < best_distance:
                    best_distance = min_distance
                    matched_name = known_face_names[best_idx]
                if min_distance < 0.6:
                    if matched_name.lower() == expected_name.lower():
                        matched = True
                        frame_to_show = frame
                        break
            if matched:
                break
        if matched:
            break

    cap.release()
    cv2.destroyAllWindows()

    if matched:
        append_log(f"✅ Match: {matched_name}")
        save_to_excel(matched_name, card_id, "Authorized")
        cv2.imwrite("temp.jpg", frame_to_show)
        show_image("temp.jpg")
        status_label.config(text=f"Access Granted: {matched_name}", fg="green")
        return "FACE_OK"
    else:
        append_log("❌ Face not matched")
        save_to_excel("Unknown", card_id, "Face Not Match", "Absent")
        status_label.config(text="Access Denied: Face Not Matched", fg="red")
        return "FACE_FAIL"

# ========== FLASK SERVER ==========
@app.route("/api/rfid")
def rfid_api():
    card_id = request.args.get("uid", "").strip()
    if not card_id:
        return "BAD_REQUEST", 400

    append_log(f"[RFID] Card Detected: {card_id}")
    evt = threading.Event()
    pending_results[card_id] = {"event": evt, "result": None}
    request_queue.put(card_id)

    processed = evt.wait(timeout=10)
    if processed:
        res = pending_results[card_id]["result"]
        del pending_results[card_id]
        return res
    else:
        del pending_results[card_id]
        append_log(f"[ERROR] Processing timeout for {card_id}")
        return "TIMEOUT", 504

# ========== QUEUE POLLING ==========
def poll_queue():
    try:
        card_id = request_queue.get_nowait()
    except queue.Empty:
        root.after(200, poll_queue)
        return

    append_log(f"[QUEUE] Processing card: {card_id}")
    result = handle_card(card_id)

    entry = pending_results.get(card_id)
    if entry:
        entry["result"] = result
        entry["event"].set()

    root.after(200, poll_queue)

# ========== REGISTER USER ==========
def register_user():
    def start_camera():
        def update_frame():
            ret, frame = cap.read()
            if ret:
                frame_resized = cv2.resize(frame, (480, 360))
                frame_rgb = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame_rgb)
                imgtk = ImageTk.PhotoImage(img)
                cam_label.imgtk = imgtk
                cam_label.configure(image=imgtk)
            cam_label.after(10, update_frame)
        update_frame()

    def capture_photo():
        name = name_entry.get().strip().lower()
        card_id = card_entry.get().strip()
        if not name or not card_id:
            messagebox.showwarning("Missing Info", "Please enter both name and card ID.")
            return
        if name == card_id:
            messagebox.showwarning("Invalid Input", "Name and Card ID cannot be the same as UID.")
            return

        ret, frame = cap.read()
        if ret:
            filename = os.path.join(KNOWN_FACES_DIR, f"{name}.jpg")
            cv2.imwrite(filename, frame)
            with open(USER_FILE, "a") as f:
                f.write(f"{card_id},{name}\n")
            messagebox.showinfo("Success", f"User {name} registered successfully!")
            load_faces()
            load_user_records()
            reg_window.destroy()
        else:
            messagebox.showerror("Capture Failed", "Unable to capture image from webcam.")

    reg_window = tk.Toplevel(root)
    reg_window.title("Register New User")
    reg_window.geometry("600x600")
    reg_window.resizable(False, False)

    ttk.Label(reg_window, text="Register New User", font=("Segoe UI", 16, "bold")).pack(pady=10)
    form_frame = ttk.Frame(reg_window)
    form_frame.pack(pady=10)
    ttk.Label(form_frame, text="Name:").grid(row=0, column=0, sticky="e")
    name_entry = ttk.Entry(form_frame, width=30)
    name_entry.grid(row=0, column=1, pady=5)
    ttk.Label(form_frame, text="Card ID:").grid(row=1, column=0, sticky="e")
    card_entry = ttk.Entry(form_frame, width=30)
    card_entry.grid(row=1, column=1, pady=5)

    cam_frame = ttk.Frame(reg_window)
    cam_frame.pack(pady=10)
    cam_label = tk.Label(cam_frame, bg="gray")
    cam_label.pack()
    ttk.Button(cam_frame, text="Capture & Register", command=capture_photo).pack(pady=10)

    cap = cv2.VideoCapture(0)
    start_camera()

    def on_close():
        cap.release()
        reg_window.destroy()
    reg_window.protocol("WM_DELETE_WINDOW", on_close)

# ========== SERVER THREAD ==========
class FlaskThread(threading.Thread):
    def run(self):
        server = make_server(HOST, PORT, app)
        server.serve_forever()

# ========== GUI CONTROLS ==========
btn_frame = tk.Frame(root, bg="#f0f2f5")
btn_frame.pack(pady=10)
ttk.Button(btn_frame, text="Register New User", command=register_user).grid(row=0, column=0, padx=10)
ttk.Button(btn_frame, text="Quit", command=root.quit).grid(row=0, column=1, padx=10)

# ========== START ==========
load_faces()
load_user_records()
append_log("✅ System Ready. Waiting for RFID scan via Wi-Fi...")
append_log(f"🌐 Connect your ESP to: http://<your_pc_ip>:{PORT}/api/rfid?uid=<card_id>")

server_thread = FlaskThread()
server_thread.daemon = True
server_thread.start()

root.after(200, poll_queue)
root.mainloop()
